import os, io, csv, re, time, requests
from datetime import datetime
from email.mime.text import MIMEText
import smtplib
from dotenv import load_dotenv
from flask import Flask, request, jsonify, send_from_directory

try:
    from openpyxl import load_workbook  # for .xlsx parsing
except Exception:
    load_workbook = None

load_dotenv()
app = Flask(__name__)

# ---------- Config ----------
PRODUCT_SHEET_URL   = os.getenv("PRODUCT_SHEET_URL", "")
SHEET_NAME          = os.getenv("SHEET_NAME", "Sheet1")
HAS_HEADER          = os.getenv("HAS_HEADER", "false").lower() == "true"
PRODUCT_COL         = int(os.getenv("PRODUCT_COL", "0"))
PRICE_COL           = int(os.getenv("PRICE_COL", "1"))
CACHE_TTL_SECONDS   = int(os.getenv("CACHE_TTL_SECONDS", "300"))

SMTP_SERVER         = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT           = int(os.getenv("SMTP_PORT", "587"))
EMAIL_ADDRESS       = os.getenv("EMAIL_ADDRESS")
EMAIL_PASSWORD      = os.getenv("EMAIL_PASSWORD")

# ---------- In-memory cache ----------
_cache = {"ts": 0, "products": []}

# ---------- Helpers ----------
def _drive_direct_download(url: str) -> str:
    """Turn shared links into direct download links (Drive/OneDrive)."""
    if "drive.google.com" in url:
        # Support both /d/<id>/ and id=<id> formats
        m = re.search(r"/d/([A-Za-z0-9_-]+)", url) or re.search(r"id=([A-Za-z0-9_-]+)", url)
        if m:
            return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
    if "1drv.ms" in url or "onedrive.live.com" in url:
        # Many OneDrive share links accept "download=1"
        if "download=1" in url:
            return url
        joiner = "&" if "?" in url else "?"
        return f"{url}{joiner}download=1"
    return url

def _parse_price(val: str):
    try:
        return float(re.sub(r"[^0-9.\-]", "", str(val)))
    except Exception:
        return None

def _fetch_products():
    now = time.time()
    if _cache["products"] and now - _cache["ts"] < CACHE_TTL_SECONDS:
        return _cache["products"]

    if not PRODUCT_SHEET_URL:
        _cache.update(ts=now, products=[])
        return []

    url = _drive_direct_download(PRODUCT_SHEET_URL)
    r = requests.get(url, timeout=30)
    r.raise_for_status()

    content_type = r.headers.get("Content-Type", "").lower()
    content = r.content
    products = []

    def push(name, price):
        name = (name or "").strip()
        p = _parse_price(price)
        if name and p is not None:
            products.append({"name": name, "price": p})

    # Try CSV by content-type or extension fallback
    if "text/csv" in content_type or url.lower().endswith(".csv"):
        f = io.StringIO(content.decode("utf-8", errors="ignore"))
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if not row:
                continue
            if i == 0 and HAS_HEADER:
                continue
            name = row[PRODUCT_COL] if len(row) > PRODUCT_COL else ""
            price = row[PRICE_COL] if len(row) > PRICE_COL else ""
            push(name, price)

    else:
        # Assume XLSX if possible
        if load_workbook is None:
            raise RuntimeError("openpyxl is not installed on server; cannot read .xlsx")
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        for i, excel_row in enumerate(ws.iter_rows(values_only=True)):
            if excel_row is None:
                continue
            if i == 0 and HAS_HEADER:
                continue
            name = excel_row[PRODUCT_COL] if len(excel_row or []) > PRODUCT_COL else ""
            price = excel_row[PRICE_COL] if len(excel_row or []) > PRICE_COL else ""
            push(name, price)

    # Cache and return
    _cache.update(ts=now, products=products)
    return products

def _send_email(subject: str, body: str):
    if not (EMAIL_ADDRESS and EMAIL_PASSWORD):
        # Email not configured; skip silently
        return
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = EMAIL_ADDRESS
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as s:
        s.starttls()
        s.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        s.send_message(msg)

def _generate_order_number():
    now = datetime.now()
    return f"THO-{now:%Y%m%d}-{int(now.timestamp())%100000:05d}"

# ---------- API ----------
@app.route("/api/products", methods=["GET"])
def api_products():
    try:
        data = _fetch_products()
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/order", methods=["POST"])
def api_order():
    data = request.json or {}
    order_no = _generate_order_number()
    # Build email body
    cust = data.get("customer", {})
    items = data.get("products", [])
    total = data.get("total", 0)

    lines = [
        f"Order No: {order_no}",
        f"Timestamp: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "Customer:",
        f"Name: {cust.get('name','')}",
        f"Phone: {cust.get('phone','')}",
        "Address:",
    ]
    for line in cust.get("address", []):
        lines.append(f"  {line}")
    lines += ["", "Items:"]
    for it in items:
        lines.append(f"- {it.get('quantity','?')} x {it.get('name','')} @ {it.get('unit_price','?')} = {it.get('line_total','?')}")
    lines += ["", f"Total: {total}"]

    body = "\n".join(lines)
    _send_email(subject=f"New Order {order_no} — Thottam Organics", body=body)
    return jsonify({"ok": True, "order_number": order_no})

# ---------- Static ----------
@app.route("/")
def serve_index():
    return send_from_directory(".", "index.html")

@app.route("/<path:filename>")
def serve_static(filename):
    return send_from_directory(".", filename)

if __name__ == "__main__":
    app.run(debug=True, port=5000)
